from dash import Input, Output, callback, no_update, callback_context, dcc, State
import dash
from pathlib import Path
from components.excel_export import create_filtered_excel_download
from dash.exceptions import PreventUpdate
import dash_bootstrap_components as dbc
from dash import html
from openpyxl import load_workbook
import traceback
import base64
import pandas as pd
import io
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill


def create_reporting_callback(app, data_dict, chart_configs):
    """
    Create a specialized callback for reporting charts that handles both
    bar chart and timeline with year range filtering.
    """

    @app.callback(
        [
            Output("reporting-bar-chart", "figure"),
            Output("timeline-bar-chart", "figure"),
        ],
        [
            Input("url", "pathname"),
            Input(
                {
                    "type": "filter-dropdown",
                    "base_id": "reporting",
                    "filter_id": "from_year",
                },
                "value",
            ),
            Input(
                {
                    "type": "filter-dropdown",
                    "base_id": "reporting",
                    "filter_id": "to_year",
                },
                "value",
            ),
        ],
        allow_duplicate=True,
    )
    def update_reporting_charts(pathname, year_from, year_to):
        if pathname != "/reporting":
            return no_update, no_update

        try:
            df = data_dict["reporting-bar"]["df"].copy()

            if year_from and year_to:
                filtered_df = df[
                    (df["reported_data_year"] >= int(year_from))
                    & (df["reported_data_year"] <= int(year_to))
                ]
            else:
                filtered_df = df

            bar_chart = chart_configs["reporting-bar"]["chart_creator"](filtered_df)
            timeline_chart = chart_configs["timeline-bar"]["chart_creator"](filtered_df)

            return bar_chart, timeline_chart

        except Exception as e:
            print(f"Error updating reporting charts: {e}")
            return no_update, no_update


def create_reporting_download_callback(app, data_dict):
    """Create download callback for reporting page"""

    @app.callback(
        Output("download-reporting-data", "data"),
        Input("btn-download-reporting-data", "n_clicks"),
        prevent_initial_call=True,
    )
    def download_reporting_data(n_clicks):
        # Get the project root directory (2 levels up from callbacks directory)
        root_dir = Path(__file__).parent.parent.parent
        input_path = root_dir / "data" / "modules.xlsx"

        # Debug print
        print(f"Looking for file at: {input_path}")
        print(f"File exists: {input_path.exists()}")

        return create_filtered_excel_download(
            input_path=input_path,
            output_filename="reporting_data.xlsx",
            sheets_to_export=[
                "Company Total Electricity Use",
                "Data Center Electricity Use ",
                "Terms_of_Data_Use",
            ],
            internal_prefix="_internal_",
            skip_rows=1,
            n_clicks=n_clicks,
        )
